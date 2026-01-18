import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font

# -------------------------------
# CONFIGURATION
# -------------------------------

DATA_PATH = "../data/student_performance_automation.csv"
SUMMARY_REPORT_PATH = "../reports/student_summary_report.xlsx"
STUDENT_REPORTS_DIR = "../reports/student_reports"

# Ensure output directories exist
os.makedirs("../reports", exist_ok=True)
os.makedirs(STUDENT_REPORTS_DIR, exist_ok=True)

# -------------------------------
# LOAD DATA
# -------------------------------

df = pd.read_csv(DATA_PATH)

# -------------------------------
# 1. CREATE SUMMARY REPORT
# -------------------------------

summary_df = df.groupby("course").agg({
    "student_id": "count",
    "attendance_percent": "mean",
    "final_score": "mean"
}).reset_index()

summary_df.rename(columns={
    "student_id": "Number of Students",
    "attendance_percent": "Average Attendance",
    "final_score": "Average Final Score"
}, inplace=True)

summary_df.to_excel(SUMMARY_REPORT_PATH, index=False)

print("✔ Summary report generated.")

# -------------------------------
# 2. CREATE INDIVIDUAL STUDENT REPORTS
# -------------------------------

for _, row in df.iterrows():
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Report"

    # Title
    ws["A1"] = "Student Performance Report"
    ws["A1"].font = Font(bold=True, size=14)

    # Student Info
    ws["A3"] = "Student Name:"
    ws["B3"] = row["student_name"]

    ws["A4"] = "Age:"
    ws["B4"] = row["age"]

    ws["A5"] = "Course:"
    ws["B5"] = row["course"]

    # Performance Data
    ws["A7"] = "Attendance Percentage:"
    ws["B7"] = row["attendance_percent"]

    ws["A8"] = "Assignment Average:"
    ws["B8"] = row["assignment_avg"]

    ws["A9"] = "Final Score:"
    ws["B9"] = row["final_score"]

    # Evaluation Logic
    if row["final_score"] >= 85:
        evaluation = "Excellent Performance"
    elif row["final_score"] >= 70:
        evaluation = "Satisfactory Performance"
    else:
        evaluation = "Needs Improvement"

    ws["A11"] = "Instructor Evaluation:"
    ws["B11"] = evaluation
    ws["B11"].font = Font(bold=True)

    # Save file
    file_name = f"{row['student_name'].replace(' ', '_')}_report.xlsx"
    file_path = os.path.join(STUDENT_REPORTS_DIR, file_name)
    wb.save(file_path)

print("✔ Individual student reports generated successfully.")
